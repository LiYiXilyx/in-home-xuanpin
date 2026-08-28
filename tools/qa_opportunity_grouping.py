import json
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

output_dir = Path(sys.argv[1] if len(sys.argv) > 1 else "outputs/catalog-grouping-2135-20260828").resolve()
workbook_path = output_dir / (sys.argv[2] if len(sys.argv) > 2 else "catalog-active-pool-2135-grouped.xlsx")
workbook = load_workbook(workbook_path, read_only=True, data_only=False)

def records(sheet_name):
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(values)]
    return headers, [list(row) for row in values if any(value not in (None, "") for value in row)]

headers, rows = records("05_细分商品明细")
index = {name: headers.index(name) for name in [
    "goods_id", "商品标题", "用户场景", "产品类型", "Level3具体细分", "相似产品簇", "排序组", "聚类依据",
    "classification_evidence", "title_evidence", "image_evidence", "evidence_conflict", "分类置信度", "销量"
]}

def text(value): return str(value or "")
def number(value):
    try: return float(value or 0)
    except (TypeError, ValueError): return 0.0
def goods(row): return text(row[index["goods_id"]]).lstrip("'")
def detail_group(row):
    return (text(row[index["用户场景"]]), text(row[index["产品类型"]]), text(row[index["Level3具体细分"]]), text(row[index["相似产品簇"]]))
def contiguous(data, key):
    closed, previous = set(), None
    for row in data:
        current = key(row)
        if current != previous:
            if current in closed: return False
            if previous is not None: closed.add(previous)
            previous = current
    return True

detail_sorted = all(detail_group(rows[i - 1]) != detail_group(rows[i]) or (-number(rows[i - 1][index["销量"]]), goods(rows[i - 1])) <= (-number(rows[i][index["销量"]]), goods(rows[i])) for i in range(1, len(rows)))
sorting_groups_match = all(text(row[index["排序组"]]) == "|".join(detail_group(row)) for row in rows)
waiting = [row for row in rows if row[index["产品类型"]] == "其它/待细分" or row[index["Level3具体细分"]] == "其它/待细分"]
waiting_clustered = [row for row in waiting if row[index["相似产品簇"]] not in (None, "", "未知")]
clusters = {text(row[index["相似产品簇"]]) for row in rows if text(row[index["相似产品簇"]]) not in ("", "未知")}

segment_headers, segment_rows = records("04_细分机会")
segment_index = {name: segment_headers.index(name) for name in ["用户场景", "产品类型", "Level3具体细分"]}
segment_keys = [(text(row[segment_index["用户场景"]]), text(row[segment_index["产品类型"]]), text(row[segment_index["Level3具体细分"]])) for row in segment_rows]

review_headers, review_rows = records("09_复核清单")
review_index = {name: review_headers.index(name) for name in ["goods_id", "相似产品簇", "分类置信度", "销量"] if name in review_headers}
def review_tail(row): return (number(row[review_index["分类置信度"]]), -number(row[review_index["销量"]]), text(row[review_index["goods_id"]]))
review_ordered = contiguous(review_rows, lambda row: text(row[review_index["相似产品簇"]])) and all(text(review_rows[i - 1][review_index["相似产品簇"]]) != text(review_rows[i][review_index["相似产品簇"]]) or review_tail(review_rows[i - 1]) <= review_tail(review_rows[i]) for i in range(1, len(review_rows)))

config = json.loads(Path("config.json").read_text(encoding="utf-8"))
database_path = Path(config["app"]["databasePath"]).resolve()
connection = sqlite3.connect(f"file:{database_path.as_posix()}?mode=ro", uri=True)
try:
    active_pool_count = int(connection.execute("SELECT product_count FROM catalog_pool_versions WHERE status='active' ORDER BY activated_at DESC,id DESC LIMIT 1").fetchone()[0])
    integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
finally:
    connection.close()

def sample(predicate):
    selected = [row for row in rows if predicate(text(row[index["相似产品簇"]]))][:10]
    return [{"goods_id": goods(row), "title": row[index["商品标题"]], "level1": row[index["用户场景"]], "level2": row[index["产品类型"]], "level3": row[index["Level3具体细分"]], "cluster": row[index["相似产品簇"]]} for row in selected]

checks = {
    "detail_rows_2135": len(rows) == 2135,
    "unique_goods_2135": len({goods(row) for row in rows}) == 2135,
    "detail_sorted": detail_sorted and sorting_groups_match,
    "same_level2_contiguous": contiguous(rows, lambda row: (row[index["用户场景"]], row[index["产品类型"]])),
    "same_level3_contiguous": contiguous(rows, lambda row: (row[index["用户场景"]], row[index["产品类型"]], row[index["Level3具体细分"]])),
    "same_similar_cluster_contiguous": contiguous(rows, lambda row: (row[index["用户场景"]], row[index["产品类型"]], row[index["Level3具体细分"]], row[index["相似产品簇"]])),
    "segments_sorted": contiguous(segment_rows, lambda row: (row[segment_index["用户场景"]], row[segment_index["产品类型"]])) and contiguous(segment_rows, lambda row: (row[segment_index["用户场景"]], row[segment_index["产品类型"]], row[segment_index["Level3具体细分"]])),
    "review_sorted": review_ordered,
    "active_pool_unchanged": active_pool_count == 2135,
    "db_integrity_ok": integrity == "ok",
}
report = {
    "pass": all(checks.values()), "workbook": str(workbook_path), "checks": checks,
    "grouping_qa": {"similar_cluster_count": len(clusters), "waiting_count": len(waiting), "waiting_clustered_count": len(waiting_clustered), "unclustered_waiting_count": len(waiting) - len(waiting_clustered)},
    "pool": {"active_pool_count": active_pool_count, "integrity": integrity},
    "samples": {"covers": sample(lambda value: value == "车罩"), "fasteners": sample(lambda value: any(word in value for word in ["螺丝", "螺栓", "螺母", "垫片", "紧固件"])), "brackets": sample(lambda value: value == "安装支架/转接件"), "bags": sample(lambda value: "包" in value)},
}
def render_preview(sheet_headers, sheet_rows, selected, target):
    positions = [sheet_headers.index(name) for name in selected]
    widths = [180, 650, 170, 190, 190, 190, 100]
    height, header_height, row_height = 48 + len(sheet_rows) * 38, 48, 38
    image = Image.new("RGB", (sum(widths), height), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", 15)
    bold = ImageFont.truetype("C:/Windows/Fonts/msyhbd.ttc", 16)
    x = 0
    for name, width in zip(selected, widths):
        draw.rectangle((x, 0, x + width, header_height), fill="#17365D")
        draw.text((x + 6, 13), name, fill="white", font=bold)
        x += width
    for row_no, row in enumerate(sheet_rows):
        y, x = header_height + row_no * row_height, 0
        fill = "#F5F8FC" if row_no % 2 else "white"
        for position, width in zip(positions, widths):
            draw.rectangle((x, y, x + width, y + row_height), fill=fill, outline="#D9E2F3")
            value = text(row[position])
            if len(value) > max(8, width // 10): value = value[:max(8, width // 10) - 1] + "…"
            draw.text((x + 6, y + 9), value, fill="#1F2937", font=font)
            x += width
    image.save(target)

(output_dir / "grouping-qa.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
render_preview(headers, rows[:24], ["goods_id", "商品标题", "用户场景", "产品类型", "Level3具体细分", "相似产品簇", "销量"], output_dir / "preview-05-grouped-detail.png")
render_preview(review_headers, review_rows[:24], ["goods_id", "标题", "产品类型", "Level3具体细分", "相似产品簇", "分类置信度", "销量"], output_dir / "preview-09-grouped-review.png")
print(json.dumps(report, ensure_ascii=False, indent=2))
sys.exit(0 if report["pass"] else 1)
